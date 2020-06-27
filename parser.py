import openpyxl
import datetime


def weekday():
    data = datetime.datetime.today().isoweekday()
    dat = {1: 'понедельник', 2: 'вторник', 3: 'среда', 4: 'четверг', 5: 'пятница', 6: 'суббота', 7: 'воскресенье'}
    print(dat[data])
    return dat[data]


def stroka():
    # d = weekday()
    d = 'среда'
    if d == 'понедельник':
        return 4, 16
    elif d == 'вторник':
        return 16, 28
    elif d == 'среда':
        return 28, 40
    elif d == 'четверг':
        return 40, 52
    elif d == 'пятница':
        return 52, 64
    elif d == 'суббота':
        return 64, 76
    elif d == 'воскресенье':
        return 0, 0


# while True:
    # date = weekday()
    date = 'среда'
    wb = openpyxl.load_workbook('./test3.xlsx')

    sheet = wb.active
    # k = 0
    # for k in range(7):
    for cell in sheet['A']:  # Нечетная неделя
        if (str(cell.value) == 'None') or (str(cell.value).lower() != str(date)):
            continue
        else:
            print('Расписание на ', date)
            c, s = stroka()
            cols = sheet.max_column
            for j in range(1, cols + 1):
                a = sheet.cell(row=3, column=j)
                if str(a.value).lower() == 'предмет':
                    string = ''
                    b = sheet.cell(row=2, column=j)
                    if str(b.value).lower()[0] == 'б':
                        print(str(b.value))
                    for i in range(c, s):
                        if i % 2 == 0:
                            predmet = sheet.cell(row=i, column=j)
                            if str(predmet.value) != 'None':
                                string = string + str(predmet.value) + ' '
                    print(string)

    # for k in range(14):
        # for cell in sheet['A']:  # Четная неделя
        #     if (str(cell.value) == 'None') or (str(cell.value).lower() != str(date)):
        #         continue
        #     else:
        #         c, s = stroka()
        #         rows = sheet.max_row
        #         cols = sheet.max_column
        #         for j in range(1, cols + 1):
        #             a = sheet.cell(row=3, column=j)
        #             if str(a.value).lower() == 'предмет':
        #                 string = ''
        #                 b = sheet.cell(row=2, column=j)
        #                 if str(b.value).lower()[0] == 'б':
        #                     print(str(b.value))
        #                 for i in range(c, s):
        #                     if i % 2 != 0:
        #                         predmet = sheet.cell(row=i, column=j)
        #                         if str(predmet.value) != 'None':
        #                             string = string + str(predmet.value) + ' '
        #                 print(string)
    #   k = 0
